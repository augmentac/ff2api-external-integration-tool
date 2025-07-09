"""
Enhanced Output Generator for LTL Tracking Integration

This module provides functionality to generate Excel templates with tracking data
integrated into the customer's original data format.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional, Any
import logging
from datetime import datetime
import os


class EnhancedOutputGenerator:
    """
    Generates enhanced Excel output with tracking data integration.
    
    This class creates Excel files that combine:
    - Original CSV data
    - API processing results
    - LTL tracking status information
    - Customer-friendly formatting
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def generate_enhanced_excel(self, 
                              original_df: pd.DataFrame,
                              processing_results: List[Dict[str, Any]],
                              tracking_results: List[Any],
                              output_path: str,
                              customer_name: str = "Customer") -> str:
        """
        Generate an enhanced Excel file with tracking data.
        
        Args:
            original_df: Original CSV data
            processing_results: Results from API processing
            tracking_results: Results from PRO number tracking
            output_path: Path to save the Excel file
            customer_name: Customer name for the template
            
        Returns:
            Path to the generated Excel file
        """
        try:
            # Create enhanced dataframe
            enhanced_df = self._create_enhanced_dataframe(
                original_df, processing_results, tracking_results
            )
            
            # Create Excel file with formatting
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{customer_name} Load Status"
            
            # Add headers and data
            self._add_data_to_worksheet(ws, enhanced_df)
            
            # Apply formatting
            self._apply_excel_formatting(ws, enhanced_df)
            
            # Add summary sheet
            self._add_summary_sheet(wb, enhanced_df, customer_name)
            
            # Save the file
            wb.save(output_path)
            
            self.logger.info(f"Enhanced Excel file generated: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error generating enhanced Excel: {e}")
            raise
    
    def _create_enhanced_dataframe(self,
                                 original_df: pd.DataFrame,
                                 processing_results: List[Dict[str, Any]],
                                 tracking_results: List[Any]) -> pd.DataFrame:
        """
        Create an enhanced dataframe with tracking data.
        
        Args:
            original_df: Original CSV data
            processing_results: API processing results
            tracking_results: Tracking results
            
        Returns:
            Enhanced DataFrame with tracking columns
        """
        # Start with original data
        enhanced_df = original_df.copy()
        
        # Add API processing status columns
        enhanced_df['API_Status'] = None
        enhanced_df['API_Load_Number'] = None
        enhanced_df['API_Error_Message'] = None
        
        # Add tracking status columns
        enhanced_df['Tracking_Status'] = None
        enhanced_df['Tracking_Location'] = None
        enhanced_df['Tracking_Event'] = None
        enhanced_df['Tracking_Timestamp'] = None
        enhanced_df['Carrier_Name'] = None
        enhanced_df['Last_Updated'] = None
        
        # Map processing results to rows
        for i, result in enumerate(processing_results):
            if i < len(enhanced_df):
                enhanced_df.at[i, 'API_Status'] = 'Success' if result.get('success', False) else 'Failed'
                enhanced_df.at[i, 'API_Load_Number'] = result.get('load_number', '')
                enhanced_df.at[i, 'API_Error_Message'] = result.get('error', '') if not result.get('success', False) else ''
        
        # Map tracking results to rows
        tracking_dict = {}
        for result in tracking_results:
            if hasattr(result, 'row_index') and result.row_index is not None:
                tracking_dict[result.row_index] = result
        
        for row_index, tracking_result in tracking_dict.items():
            if row_index < len(enhanced_df):
                enhanced_df.at[row_index, 'Tracking_Status'] = tracking_result.tracking_status or 'Unknown'
                enhanced_df.at[row_index, 'Tracking_Location'] = tracking_result.tracking_location or ''
                enhanced_df.at[row_index, 'Tracking_Event'] = tracking_result.tracking_event or ''
                enhanced_df.at[row_index, 'Tracking_Timestamp'] = tracking_result.tracking_timestamp or ''
                enhanced_df.at[row_index, 'Carrier_Name'] = tracking_result.carrier_name or 'Unknown'
                enhanced_df.at[row_index, 'Last_Updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        return enhanced_df
    
    def _add_data_to_worksheet(self, ws, df: pd.DataFrame):
        """Add dataframe data to worksheet"""
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    
    def _apply_excel_formatting(self, ws, df: pd.DataFrame):
        """Apply professional formatting to the Excel worksheet"""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply conditional formatting for status columns
        self._apply_conditional_formatting(ws, df)
    
    def _apply_conditional_formatting(self, ws, df: pd.DataFrame):
        """Apply conditional formatting for status columns"""
        # Find status columns
        api_status_col = None
        tracking_status_col = None
        
        for i, col in enumerate(df.columns):
            if col == 'API_Status':
                api_status_col = i + 1
            elif col == 'Tracking_Status':
                tracking_status_col = i + 1
        
        # Color code API status
        if api_status_col:
            for row in range(2, len(df) + 2):
                cell = ws.cell(row=row, column=api_status_col)
                if cell.value == 'Success':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif cell.value == 'Failed':
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Color code tracking status
        if tracking_status_col:
            for row in range(2, len(df) + 2):
                cell = ws.cell(row=row, column=tracking_status_col)
                status = str(cell.value).lower() if cell.value else ''
                
                if any(keyword in status for keyword in ['delivered', 'complete']):
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif any(keyword in status for keyword in ['in transit', 'departed', 'picked up']):
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif any(keyword in status for keyword in ['delayed', 'exception', 'problem']):
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    def _add_summary_sheet(self, wb, df: pd.DataFrame, customer_name: str):
        """Add a summary sheet with key metrics"""
        ws = wb.create_sheet(title="Summary")
        
        # Title
        ws['A1'] = f"{customer_name} Load Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Generation timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:D2')
        
        # Summary metrics
        total_loads = len(df)
        api_success = len(df[df['API_Status'] == 'Success'])
        api_failed = len(df[df['API_Status'] == 'Failed'])
        tracking_available = len(df[df['Tracking_Status'].notna() & (df['Tracking_Status'] != 'Unknown')])
        
        # Create metrics table
        metrics = [
            ['Metric', 'Value'],
            ['Total Loads', total_loads],
            ['API Success', api_success],
            ['API Failed', api_failed],
            ['Tracking Available', tracking_available],
            ['Success Rate', f"{(api_success/total_loads*100):.1f}%" if total_loads > 0 else "0%"],
            ['Tracking Rate', f"{(tracking_available/total_loads*100):.1f}%" if total_loads > 0 else "0%"]
        ]
        
        # Add metrics to worksheet
        for row_idx, row_data in enumerate(metrics, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_tracking_only_report(self, 
                                    tracking_results: List[Any],
                                    output_path: str,
                                    customer_name: str = "Customer") -> str:
        """
        Generate a tracking-only report for customers.
        
        Args:
            tracking_results: List of tracking results
            output_path: Path to save the report
            customer_name: Customer name for the report
            
        Returns:
            Path to the generated report
        """
        try:
            # Create dataframe from tracking results
            tracking_data = []
            for result in tracking_results:
                if hasattr(result, 'pro_number'):
                    tracking_data.append({
                        'PRO_Number': result.pro_number,
                        'Carrier': result.carrier_name,
                        'Status': result.tracking_status or 'Unknown',
                        'Location': result.tracking_location or '',
                        'Latest_Event': result.tracking_event or '',
                        'Timestamp': result.tracking_timestamp or '',
                        'Last_Updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
            
            if not tracking_data:
                raise ValueError("No tracking data available")
            
            df = pd.DataFrame(tracking_data)
            
            # Create Excel file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{customer_name} Tracking Report"
            
            # Add data
            self._add_data_to_worksheet(ws, df)
            
            # Apply formatting
            self._apply_excel_formatting(ws, df)
            
            # Save file
            wb.save(output_path)
            
            self.logger.info(f"Tracking report generated: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error generating tracking report: {e}")
            raise


def generate_customer_output(upload_id: int,
                           db_manager,
                           customer_name: str = "Customer",
                           output_dir: str = "data/outputs") -> Optional[str]:
    """
    Generate customer output file with tracking data.
    
    Args:
        upload_id: Upload history ID
        db_manager: Database manager instance
        customer_name: Customer name
        output_dir: Output directory
        
    Returns:
        Path to generated file or None if error
    """
    try:
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Get tracking results from database
        tracking_results = db_manager.get_tracking_results_for_upload(upload_id)
        
        if not tracking_results:
            logging.warning(f"No tracking results found for upload {upload_id}")
            return None
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{customer_name}_tracking_report_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, filename)
        
        # Generate report
        generator = EnhancedOutputGenerator()
        return generator.generate_tracking_only_report(
            tracking_results, output_path, customer_name
        )
        
    except Exception as e:
        logging.error(f"Error generating customer output: {e}")
        return None 